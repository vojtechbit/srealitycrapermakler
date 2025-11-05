#!/usr/bin/env python3
"""Unified CLI for scraping real-estate agent contacts."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable, List, Sequence

import pandas as pd

from scrapers import get_scraper, list_scrapers
from scrapers.base import BaseScraper, ScraperResult, merge_results


def _available_slugs() -> List[str]:
    return sorted(scraper.slug for scraper in list_scrapers())


def _prompt_for_platform() -> List[str]:
    scrapers = list_scrapers()
    print("Dostupné platformy:")
    for scraper in scrapers:
        print(f"  - {scraper.slug:15s} {scraper.name}")
    selected = input("Zadejte platformy oddělené čárkou (např. 'sreality,linkedin'): ")
    slugs = [slug.strip() for slug in selected.split(",") if slug.strip()]
    return slugs or ["sreality"]


def _prompt_for_sreality_params() -> dict:
    """Interaktivně se zeptá na parametry pro Sreality.cz"""
    print("\n" + "="*60)
    print("NASTAVENÍ PARAMETRŮ PRO SREALITY.CZ")
    print("="*60)

    # Kategorie nemovitosti
    print("\n📋 Typ nemovitosti (můžeš vybrat víc oddělených čárkou):")
    print("  1 - Byty")
    print("  2 - Domy")
    print("  3 - Pozemky")
    print("  4 - Komerční")
    print("  5 - Ostatní")
    print("  Příklad: '1,2' = Byty a Domy")

    while True:
        category_input = input("Vyber typ nemovitosti (1-5) [1]: ").strip() or "1"
        categories = [c.strip() for c in category_input.split(",")]
        if all(c in ["1", "2", "3", "4", "5"] for c in categories):
            category_main_list = [int(c) for c in categories]
            break
        print("❌ Neplatná volba, zadej čísla 1-5 oddělená čárkou (např. '1,2')")

    # Typ inzerátu
    print("\n📋 Typ inzerátu (můžeš vybrat víc oddělených čárkou):")
    print("  1 - Prodej")
    print("  2 - Pronájem")
    print("  3 - Dražby")
    print("  Příklad: '1,2' = Prodej a Pronájem")

    while True:
        type_input = input("Vyber typ inzerátu (1-3) [1]: ").strip() or "1"
        types = [t.strip() for t in type_input.split(",")]
        if all(t in ["1", "2", "3"] for t in types):
            category_type_list = [int(t) for t in types]
            break
        print("❌ Neplatná volba, zadej čísla 1-3 oddělená čárkou (např. '1,2')")

    # Kraj
    print("\n📋 Kraj (můžeš vybrat víc oddělených čárkou):")
    print("  10 - Praha")
    print("  11 - Středočeský")
    print("  12 - Jihočeský")
    print("  13 - Plzeňský")
    print("  14 - Karlovarský")
    print("  15 - Ústecký")
    print("  16 - Liberecký")
    print("  17 - Královéhradecký")
    print("  18 - Pardubický")
    print("  19 - Vysočina")
    print("  20 - Jihomoravský")
    print("  21 - Olomoucký")
    print("  22 - Zlínský")
    print("  23 - Moravskoslezský")
    print("  Příklad: '10,20' = Praha a Jihomoravský")

    locality_input = input("Vyber kraje (10-23) oddělené čárkou nebo Enter pro celou ČR: ").strip()
    if locality_input:
        localities = [loc.strip() for loc in locality_input.split(",")]
        try:
            locality_list = [int(loc) for loc in localities if loc]
            # Validace
            if not all(10 <= loc <= 23 for loc in locality_list):
                print("⚠️  Některé kódy krajů jsou mimo rozsah 10-23, budu je ignorovat")
                locality_list = [loc for loc in locality_list if 10 <= loc <= 23]
        except ValueError:
            print("⚠️  Neplatné číslo kraje, použiji celou ČR")
            locality_list = [None]
    else:
        locality_list = [None]  # Celá ČR

    # Maximální počet stránek
    print("\n📋 Rozsah scrapování:")
    print("  0 - VŠECHNY stránky (může trvat hodiny!)")
    print("  1-10 - Rychlý test")
    print("  20-50 - Střední rozsah")
    print("  100+ - Velký rozsah")

    while True:
        max_pages = input("Maximální počet stránek [5]: ").strip() or "5"
        try:
            max_pages = int(max_pages)
            if max_pages < 0:
                print("❌ Zadej kladné číslo nebo 0")
                continue
            break
        except ValueError:
            print("❌ Zadej platné číslo")

    # Full scan
    full_scan = False
    if max_pages == 0:
        full_scan = True
        max_pages = None

    # Detaily
    print("\n📋 Stahovat detaily inzerátů?")
    print("  y - Ano (přesnější kontakty, ale POMALEJŠÍ - cca 2-3 sekundy na inzerát)")
    print("  n - Ne (rychlejší, ale méně kontaktů)")

    fetch_details = input("Stahovat detaily? (y/n) [y]: ").strip().lower() or "y"
    fetch_details = fetch_details == "y"

    # Vypočítej celkový počet kombinací
    total_combinations = len(category_main_list) * len(category_type_list) * len(locality_list)

    print("\n" + "="*60)
    print("SOUHRN:")
    print("="*60)
    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
    region_names = {
        10: "Praha", 11: "Středočeský", 12: "Jihočeský", 13: "Plzeňský",
        14: "Karlovarský", 15: "Ústecký", 16: "Liberecký", 17: "Královéhradecký",
        18: "Pardubický", 19: "Vysočina", 20: "Jihomoravský", 21: "Olomoucký",
        22: "Zlínský", 23: "Moravskoslezský"
    }

    category_names_str = ", ".join(category_names.get(c, 'Neznámý') for c in category_main_list)
    type_names_str = ", ".join(type_names.get(t, 'Neznámý') for t in category_type_list)
    region_names_str = ", ".join(region_names.get(loc, 'Celá ČR') for loc in locality_list)

    print(f"📌 Typ nemovitosti: {category_names_str}")
    print(f"📌 Typ inzerátu: {type_names_str}")
    print(f"📌 Kraje: {region_names_str}")
    print(f"📌 Max. stránek: {'VŠECHNY' if full_scan else max_pages}")
    print(f"📌 Detaily: {'Ano' if fetch_details else 'Ne'}")
    print(f"\n⚠️  Celkem kombinací k scrapování: {total_combinations}")
    if total_combinations > 1:
        print(f"    (Scraper poběží {total_combinations}x a výsledky sloučí)")
    print("="*60)

    confirm = input("\nPokračovat? (y/n) [y]: ").strip().lower() or "y"
    if confirm != "y":
        print("❌ Zrušeno uživatelem")
        sys.exit(0)

    return {
        "category_main_list": category_main_list,
        "category_type_list": category_type_list,
        "locality_list": locality_list,
        "max_pages": max_pages,
        "full_scan": full_scan,
        "fetch_details": fetch_details,
    }


def _parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--platform",
        "-p",
        action="append",
        dest="platforms",
        help="Vybrané platformy (slug). Lze zadat vícekrát.",
    )
    parser.add_argument(
        "--all-platforms",
        action="store_true",
        help="Spustí scraping na všech dostupných platformách.",
    )
    parser.add_argument(
        "--full-scan",
        action="store_true",
        help="Pokud platforma podporuje, projde všechny stránky (může trvat hodiny).",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        help="Maximální počet stránek pro platformy, které stránkují výsledky.",
    )
    parser.add_argument(
        "--category-main",
        type=int,
        default=1,
        help="Kategorie nemovitostí pro Sreality (1=Byty, 2=Domy, 3=Pozemky, 4=Komerční, 5=Ostatní).",
    )
    parser.add_argument(
        "--category-type",
        type=int,
        default=1,
        help="Typ nabídky pro Sreality (1=Prodej, 2=Pronájem, 3=Dražby).",
    )
    parser.add_argument(
        "--locality",
        type=int,
        help="ID regionu pro Sreality (např. 10 = Praha).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Cílový soubor (Excel .xlsx). Bez zadání se jen vypíše souhrn.",
    )
    parser.add_argument(
        "--list",
        action="store_true",
        help="Pouze vypíše dostupné platformy a skončí.",
    )
    parser.add_argument(
        "--prompt",
        action="store_true",
        help="Interaktivně se zeptá na výběr platformy, pokud není zadána.",
    )
    return parser.parse_args(argv)


def _validate_platforms(platforms: Iterable[str]) -> List[str]:
    available = set(_available_slugs())
    invalid = [slug for slug in platforms if slug not in available]
    if invalid:
        raise SystemExit(f"Neznámé platformy: {', '.join(invalid)}")
    return list(dict.fromkeys(platforms))


def _save_to_excel(result: ScraperResult, output: Path) -> None:
    """Uloží výsledky do Excelu s hyperlinky a formátováním."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(result.records)

    # Nejdřív ulož běžně
    df.to_excel(output, index=False, engine="openpyxl")

    # Pak načti a přidej hyperlinky + formátování
    wb = load_workbook(output)
    ws = wb.active

    # Najdi sloupce s linky
    headers = [cell.value for cell in ws[1]]
    link_columns = []

    for idx, header in enumerate(headers, 1):
        if header in ["profil_maklere", "odkazy", "profil_url"]:
            link_columns.append((idx, header))

    # Přidej hyperlinky
    for row_idx in range(2, ws.max_row + 1):  # Začni od řádku 2 (přeskoč hlavičku)
        for col_idx, col_name in link_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value

            if url and isinstance(url, str) and url.startswith("http"):
                # Zkrať text pro odkazy (pokud je to seznam URL)
                if col_name == "odkazy" and "," in url:
                    # Více odkazů - zobraz jen "Více odkazů"
                    urls = [u.strip() for u in url.split(",") if u.strip()]
                    first_url = urls[0]
                    cell.hyperlink = first_url
                    cell.value = f"Zobrazit ({len(urls)} inzerátů)"
                    cell.font = Font(color="0000FF", underline="single")
                else:
                    # Jeden odkaz
                    cell.hyperlink = url
                    # Zkrať zobrazení
                    if col_name == "profil_maklere":
                        cell.value = "Profil makléře"
                    cell.font = Font(color="0000FF", underline="single")

    # Automatická šířka sloupců
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Nastav šířku (s limitem)
        adjusted_width = min(max_length + 2, 60)  # Max 60 znaků
        ws.column_dimensions[column_letter].width = adjusted_width

    # Ulož změny
    wb.save(output)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv or sys.argv[1:])

    if args.list:
        print("Dostupné platformy:")
        for scraper in list_scrapers():
            print(f"- {scraper.slug:15s} {scraper.name:30s} | {scraper.description}")
        return 0

    if args.all_platforms:
        platforms = _available_slugs()
    elif args.platforms:
        platforms = _validate_platforms(args.platforms)
    elif args.prompt:
        platforms = _validate_platforms(_prompt_for_platform())
    else:
        # Default behaviour: ask a simple question (backwards compatible).
        platforms = _validate_platforms(_prompt_for_platform())

    # If using --prompt and sreality is selected, ask for parameters interactively
    sreality_params = None
    if args.prompt and "sreality" in platforms:
        sreality_params = _prompt_for_sreality_params()

    print("\nSpouštím scraping pro:")
    for slug in platforms:
        scraper = get_scraper(slug)
        print(f"- {scraper.name} ({slug})")

    results: List[ScraperResult] = []

    for slug in platforms:
        scraper = get_scraper(slug)

        # Use interactive params for sreality if available, otherwise use args
        if slug == "sreality" and sreality_params:
            # Multiple combinations - loop through all
            category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
            type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
            region_names = {
                10: "Praha", 11: "Středočeský", 12: "Jihočeský", 13: "Plzeňský",
                14: "Karlovarský", 15: "Ústecký", 16: "Liberecký", 17: "Královéhradecký",
                18: "Pardubický", 19: "Vysočina", 20: "Jihomoravský", 21: "Olomoucký",
                22: "Zlínský", 23: "Moravskoslezský"
            }

            total_combinations = (
                len(sreality_params["category_main_list"]) *
                len(sreality_params["category_type_list"]) *
                len(sreality_params["locality_list"])
            )
            current_combo = 0

            for category_main in sreality_params["category_main_list"]:
                for category_type in sreality_params["category_type_list"]:
                    for locality in sreality_params["locality_list"]:
                        current_combo += 1

                        print("\n" + "="*60)
                        print(f"Platforma: {scraper.name} ({slug})")
                        print(f"Kombinace {current_combo}/{total_combinations}")
                        print(f"  • Typ: {category_names.get(category_main, 'Neznámý')}")
                        print(f"  • Inzerát: {type_names.get(category_type, 'Neznámý')}")
                        print(f"  • Kraj: {region_names.get(locality, 'Celá ČR')}")
                        print("="*60)

                        kwargs = {
                            "category_main": category_main,
                            "category_type": category_type,
                            "locality_region_id": locality,
                            "fetch_details": sreality_params["fetch_details"],
                        }
                        max_pages = sreality_params["max_pages"]
                        full_scan = sreality_params["full_scan"]

                        result = scraper.scrape(
                            max_pages=max_pages,
                            full_scan=full_scan,
                            **kwargs,
                        )
                        if result.records:
                            print(f"✓ {len(result.records)} záznamů")
                        if result.warnings:
                            print("⚠️  Varování:")
                            for warning in result.warnings:
                                print(f"   - {warning}")
                        if result.errors:
                            print("❌ Chyby:")
                            for error in result.errors:
                                print(f"   - {error}")
                        results.append(result)
        else:
            # Single run with command-line args
            kwargs = {
                "category_main": args.category_main,
                "category_type": args.category_type,
                "locality_region_id": args.locality,
            }
            max_pages = args.max_pages
            full_scan = args.full_scan

            print("\n==============================")
            print(f"Platforma: {scraper.name} ({slug})")
            print(f"Popis: {scraper.description}")
            print(f"Rate-limit: {scraper.rate_limit_info}")
            if full_scan and not scraper.supports_full_scan:
                print("⚠️  Platforma nepodporuje plný průchod, použiji dostupný režim.")
            result = scraper.scrape(
                max_pages=max_pages,
                full_scan=full_scan,
                **kwargs,
            )
            if result.records:
                print(f"✓ {len(result.records)} záznamů")
            if result.warnings:
                print("⚠️  Varování:")
                for warning in result.warnings:
                    print(f"   - {warning}")
            if result.errors:
                print("❌ Chyby:")
                for error in result.errors:
                    print(f"   - {error}")
            results.append(result)

    merged = merge_results(results)
    print("\n==============================")
    print(f"Celkem nalezeno {len(merged.records)} unikátních záznamů.")

    # Auto-generate output filename if using --prompt and no output specified
    output_path = args.output
    if not output_path and args.prompt and merged.records:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(f"data/makleri_{timestamp}.xlsx")
        print(f"\nAutomaticky vytvořen název souboru: {output_path}")

    if output_path:
        _save_to_excel(merged, output_path)
        print(f"✅ Data uložena do {output_path}")
    else:
        print("\n⚠️  Nezadal jsi --output, data nejsou uložena do souboru.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
