#!/usr/bin/env python3
"""
🚀 SUPER RYCHLÝ scraper makléřů s využitím company API

Logika:
1. Projde inzeráty, agreguje podle company_id (rychlé)
2. Pro každou company stáhne seznam makléřů z API (rychlé!)
3. Vytvoří hierarchický Excel: Company → Makléři

Rychlost: 4× rychlejší než předchozí verze!
"""

import argparse
import sys
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scrapers.sreality import SrealityScraper


def slugify_company_name(name):
    """Převede název company na URL-friendly slug."""
    if not name or not isinstance(name, str):
        return "company"
    normalized = unicodedata.normalize("NFKD", name)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.lower()
    ascii_value = re.sub(r"[^a-z0-9]+", "-", ascii_value)
    ascii_value = re.sub(r"-+", "-", ascii_value)
    return ascii_value.strip("-") or "company"


def scrape_agents_fast_combined(
    scraper,
    combinations,  # List of (category_main, category_type, locality) tuples
    max_pages,
    full_scan,
):
    """
    Super rychlý scraping s deduplikací companies napříč kombinacemi.

    FÁZE 1: Agreguj companies ze VŠECH kombinací
    FÁZE 2: Deduplikuj (každá company jen jednou)
    FÁZE 3: Volej sellers API jen pro unikátní companies
    """

    print(f"🔍 FÁZE 1: Agregace companies ze všech kombinací...")

    if full_scan:
        max_pages = None

    limit = max_pages if max_pages is not None else None

    # Sdílený dictionary pro VŠECHNY kombinace!
    all_companies = defaultdict(lambda: {
        "company_id": None,
        "company_name": None,
        "total_estates": 0,
        "localities": set(),
        "category_breakdown": defaultdict(int),
    })

    total_listings_all = 0
    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}

    # FÁZE 1: Projdi VŠECHNY kombinace a agreguj do jednoho dictionary
    for combo_idx, (category_main, category_type, locality_region_id) in enumerate(combinations, 1):
        print(f"\n   Kombinace {combo_idx}/{len(combinations)}: {category_names.get(category_main)} / {type_names.get(category_type)}")

        page = 1

        while True:
            if limit is not None and page > limit:
                break

            params = {
                "category_main_cb": category_main,
                "category_type_cb": category_type,
                "page": page,
                "per_page": 60,
            }

            if locality_region_id is not None:
                params["locality_region_id"] = locality_region_id

            payload = scraper._request(scraper._config.api_url, params=params)
            if not payload:
                print(f"      ⚠️  Chyba při stahování stránky {page}")
                break

            estates = payload.get("_embedded", {}).get("estates", [])
            if not estates:
                break

            # Počítadla
            new_companies = 0
            existing_companies = 0

            for estate in estates:
                total_listings_all += 1

                embedded = estate.get("_embedded", {})
                company = embedded.get("company", {})

                if not company:
                    continue

                company_id = company.get("id")
                if not company_id:
                    continue

                company_id = str(company_id)

                # Kontrola, jestli je company nová (napříč VŠEMI kombinacemi!)
                comp = all_companies[company_id]

                if comp["company_id"] is None:
                    comp["company_id"] = company_id
                    comp["company_name"] = company.get("name")
                    new_companies += 1
                else:
                    existing_companies += 1

                comp["total_estates"] += 1

                # Lokalita
                locality = estate.get("locality", "")
                if locality:
                    comp["localities"].add(locality)

                # Kategorie
                seo = estate.get("seo", {}) if isinstance(estate.get("seo"), dict) else {}
                cat_main = seo.get("category_main_cb") or category_main
                cat_type = seo.get("category_type_cb") or category_type
                key = (cat_main, cat_type)
                comp["category_breakdown"][key] += 1

            # Výpis
            print(f"      Stránka {page}: {len(estates)} inzerátů", end="")
            if new_companies > 0 or existing_companies > 0:
                print(f" (Nové RK: {new_companies}, Existující: {existing_companies})", end="")
            print()

            result_size = payload.get("result_size", 0)
            if (page * 60) >= result_size:
                break

            page += 1
            scraper._delay()

    print(f"\n✅ Zpracováno {total_listings_all} inzerátů celkem")
    print(f"✅ Nalezeno {len(all_companies)} UNIKÁTNÍCH realitních kanceláří")

    # FÁZE 2: Volej sellers API jen pro UNIKÁTNÍ companies
    print(f"\n🔍 FÁZE 2: Stahuji seznam makléřů (jen pro unikátní RK)...")

    all_records = []

    for idx, (company_id, comp) in enumerate(all_companies.items(), 1):
        # Stáhnout VŠECHNY makléře (může být více stránek!)
        all_sellers = []
        page = 1

        while True:
            company_url = f"{scraper._config.base_url}/api/cs/v2/companies/{company_id}"
            params = {"page": page} if page > 1 else None
            company_data = scraper._request(company_url, params=params)

            if not company_data:
                print(f"   ⚠️  Chyba při stahování company {company_id}")
                break

            # Získej seznam makléřů
            embedded = company_data.get("_embedded", {})
            sellers_data = embedded.get("sellers", {})

            if isinstance(sellers_data, dict):
                result_size = sellers_data.get("result_size", 0)
                per_page = sellers_data.get("per_page", 20)
                sellers_list = sellers_data.get("sellers", [])
            else:
                sellers_list = []
                result_size = 0
                per_page = 20

            if not sellers_list:
                break

            all_sellers.extend(sellers_list)

            # Kontrola, jestli jsou další stránky
            if (page * per_page) >= result_size:
                break

            page += 1
            scraper._delay()

        if not all_sellers:
            print(f"   ⚠️  Company {comp['company_name']}: žádní makléři")
            continue

        # Výpis
        if page > 1:
            print(f"   {idx}/{len(all_companies)}: {comp['company_name']} - {len(all_sellers)} makléřů ({page} stránek)")
        else:
            print(f"   {idx}/{len(all_companies)}: {comp['company_name']} - {len(all_sellers)} makléřů")

        # Lokalita
        localities_list = list(comp["localities"])
        if localities_list:
            locality = localities_list[0]
            parts = [p.strip() for p in locality.split(",")]
            mesto = parts[0] if parts else ""
            kraj = parts[-1] if len(parts) > 1 else ""
        else:
            mesto = ""
            kraj = ""

        # Rozložení
        breakdown_items = []
        for (cat, typ), count in sorted(comp["category_breakdown"].items(), key=lambda x: -x[1]):
            cat_name = category_names.get(cat, f"Kategorie {cat}")
            typ_name = type_names.get(typ, f"Typ {typ}")
            breakdown_items.append(f"{cat_name}/{typ_name}: {count}")
        rozlozeni = ", ".join(breakdown_items) if breakdown_items else ""

        company_slug = slugify_company_name(comp["company_name"])

        # Company řádek
        all_records.append({
            "typ_radku": "COMPANY",
            "zdroj": "Sreality.cz",
            "realitni_kancelar": comp["company_name"],
            "jmeno_maklere": "",
            "telefon": "",
            "email": "",
            "kraj": kraj,
            "mesto": mesto,
            "profil_url": "",
            "pocet_inzeratu": comp["total_estates"],
            "rozlozeni_inzeratu": rozlozeni,
        })

        # Makléři
        for seller in all_sellers:
            seller_id = seller.get("id")
            seller_name = seller.get("name", "")

            phones = seller.get("phones", [])
            phone = ""
            if phones and isinstance(phones, list):
                first_phone = phones[0]
                if isinstance(first_phone, dict):
                    phone = first_phone.get("number", "")

            email = seller.get("email", "")
            profile_url = f"https://www.sreality.cz/adresar/{company_slug}/{company_id}/makleri/{seller_id}"

            all_records.append({
                "typ_radku": "AGENT",
                "zdroj": "",
                "realitni_kancelar": "",
                "jmeno_maklere": seller_name,
                "telefon": phone,
                "email": email,
                "kraj": "",
                "mesto": "",
                "profil_url": profile_url,
                "pocet_inzeratu": "",
                "rozlozeni_inzeratu": "",
            })

        scraper._delay()

    print(f"\n✅ Stahování dokončeno")

    return all_records


def scrape_agents_fast(
    scraper,
    category_main,
    category_type,
    locality_region_id,
    max_pages,
    full_scan,
):
    """Super rychlý scraping pomocí company API (single combination)."""

    print(f"🔍 FÁZE 1: Agregace podle company...")

    if full_scan:
        max_pages = None

    limit = max_pages if max_pages is not None else None

    # Agregace podle company_id
    companies = defaultdict(lambda: {
        "company_id": None,
        "company_name": None,
        "total_estates": 0,
        "localities": set(),  # Různé lokality
        "category_breakdown": defaultdict(int),
    })

    page = 1
    total_listings = 0

    # FÁZE 1: Projdi inzeráty a agreguj podle company
    while True:
        if limit is not None and page > limit:
            break

        params = {
            "category_main_cb": category_main,
            "category_type_cb": category_type,
            "page": page,
            "per_page": 60,
        }

        if locality_region_id is not None:
            params["locality_region_id"] = locality_region_id

        payload = scraper._request(scraper._config.api_url, params=params)
        if not payload:
            print(f"⚠️  Chyba při stahování stránky {page}")
            break

        estates = payload.get("_embedded", {}).get("estates", [])
        if not estates:
            break

        # Počítadla pro tuto stránku
        new_companies = 0
        existing_companies = 0

        for estate in estates:
            total_listings += 1

            embedded = estate.get("_embedded", {})
            company = embedded.get("company", {})

            if not company:
                continue

            company_id = company.get("id")
            if not company_id:
                continue

            company_id = str(company_id)

            # Kontrola, jestli je company nová
            is_new = company_id not in companies or companies[company_id]["company_id"] is None

            comp = companies[company_id]

            if comp["company_id"] is None:
                comp["company_id"] = company_id
                comp["company_name"] = company.get("name")
                new_companies += 1
            else:
                existing_companies += 1

            comp["total_estates"] += 1

            # Lokalita
            locality = estate.get("locality", "")
            if locality:
                comp["localities"].add(locality)

            # Kategorie
            seo = estate.get("seo", {}) if isinstance(estate.get("seo"), dict) else {}
            cat_main = seo.get("category_main_cb") or category_main
            cat_type = seo.get("category_type_cb") or category_type
            key = (cat_main, cat_type)
            comp["category_breakdown"][key] += 1

        # Výpis statistik pro tuto stránku
        print(f"   Stránka {page}: {len(estates)} inzerátů")
        if new_companies > 0 or existing_companies > 0:
            print(f"      → Nové RK: {new_companies}, Existující RK: {existing_companies}")

        result_size = payload.get("result_size", 0)
        if (page * 60) >= result_size:
            break

        page += 1
        scraper._delay()

    print(f"\n✅ Zpracováno {total_listings} inzerátů")
    print(f"✅ Nalezeno {len(companies)} realitních kanceláří")

    # FÁZE 2: Pro každou company stáhni seznam makléřů (s paginací!)
    print(f"\n🔍 FÁZE 2: Stahuji seznam makléřů z company API...")

    all_records = []

    for idx, (company_id, comp) in enumerate(companies.items(), 1):
        # Stáhnout VŠECHNY makléře (může být více stránek!)
        all_sellers = []
        page = 1

        while True:
            company_url = f"{scraper._config.base_url}/api/cs/v2/companies/{company_id}"
            params = {"page": page} if page > 1 else None
            company_data = scraper._request(company_url, params=params)

            if not company_data:
                print(f"   ⚠️  Chyba při stahování company {company_id}")
                break

            # Získej seznam makléřů
            embedded = company_data.get("_embedded", {})
            sellers_data = embedded.get("sellers", {})

            if isinstance(sellers_data, dict):
                result_size = sellers_data.get("result_size", 0)
                per_page = sellers_data.get("per_page", 20)
                sellers_list = sellers_data.get("sellers", [])
            else:
                sellers_list = []
                result_size = 0
                per_page = 20

            if not sellers_list:
                break

            all_sellers.extend(sellers_list)

            # Kontrola, jestli jsou další stránky
            if (page * per_page) >= result_size:
                break

            page += 1
            scraper._delay()  # Delay mezi stránkami

        if not all_sellers:
            print(f"   ⚠️  Company {comp['company_name']}: žádní makléři")
            continue

        # Pokud bylo více stránek, ukaž to
        if page > 1:
            print(f"   {idx}/{len(companies)}: {comp['company_name']} - {len(all_sellers)} makléřů ({page} stránek)")
        else:
            print(f"   {idx}/{len(companies)}: {comp['company_name']} - {len(all_sellers)} makléřů")

        # Lokalita - vezmi nejčastější
        localities_list = list(comp["localities"])
        if localities_list:
            # Vezmi první lokalitu a parse kraj/město
            locality = localities_list[0]
            parts = [p.strip() for p in locality.split(",")]
            mesto = parts[0] if parts else ""
            kraj = parts[-1] if len(parts) > 1 else ""
        else:
            mesto = ""
            kraj = ""

        # Vytvoř rozložení inzerátů
        category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
        type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
        breakdown_items = []
        for (cat, typ), count in sorted(comp["category_breakdown"].items(), key=lambda x: -x[1]):
            cat_name = category_names.get(cat, f"Kategorie {cat}")
            typ_name = type_names.get(typ, f"Typ {typ}")
            breakdown_items.append(f"{cat_name}/{typ_name}: {count}")
        rozlozeni = ", ".join(breakdown_items) if breakdown_items else ""

        # Company řádek (hlavička)
        company_slug = slugify_company_name(comp["company_name"])

        all_records.append({
            "typ_radku": "COMPANY",  # Speciální typ pro formátování
            "zdroj": "Sreality.cz",
            "realitni_kancelar": comp["company_name"],
            "jmeno_maklere": "",
            "telefon": "",
            "email": "",
            "kraj": kraj,
            "mesto": mesto,
            "profil_url": "",
            "pocet_inzeratu": comp["total_estates"],
            "rozlozeni_inzeratu": rozlozeni,
        })

        # Makléři pod company
        for seller in all_sellers:
            seller_id = seller.get("id")
            seller_name = seller.get("name", "")

            # Telefon - vezmi první
            phones = seller.get("phones", [])
            phone = ""
            if phones and isinstance(phones, list):
                first_phone = phones[0]
                if isinstance(first_phone, dict):
                    phone = first_phone.get("number", "")

            email = seller.get("email", "")

            # URL profilu
            profile_url = f"https://www.sreality.cz/adresar/{company_slug}/{company_id}/makleri/{seller_id}"

            all_records.append({
                "typ_radku": "AGENT",  # Makléř
                "zdroj": "",
                "realitni_kancelar": "",  # Prázdné, je pod hlavičkou
                "jmeno_maklere": seller_name,
                "telefon": phone,
                "email": email,
                "kraj": "",
                "mesto": "",
                "profil_url": profile_url,
                "pocet_inzeratu": "",
                "rozlozeni_inzeratu": "",
            })

        scraper._delay()

    print(f"\n✅ Stahuji dokončeno")

    return all_records


def save_to_excel_hierarchical(records, output_path):
    """Uloží do Excelu s hierarchickým formátováním."""
    if not records:
        print("⚠️  Žádné záznamy")
        return

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Odstraň typ_radku pro export
    export_records = []
    for rec in records:
        export_rec = {k: v for k, v in rec.items() if k != "typ_radku"}
        export_records.append(export_rec)

    df = pd.DataFrame(export_records)
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Formátování
    wb = load_workbook(output_path)
    ws = wb.active

    # Barvy
    company_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")  # Světle modrá
    company_font = Font(bold=True, size=12)

    # Najdi sloupec profil_url
    headers = [cell.value for cell in ws[1]]
    profil_col = None
    for idx, header in enumerate(headers, 1):
        if header == "profil_url":
            profil_col = idx
            break

    # Formátuj řádky
    for row_idx in range(2, ws.max_row + 1):
        typ_radku = records[row_idx - 2].get("typ_radku")

        if typ_radku == "COMPANY":
            # Company řádek - zvýrazni
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = company_fill
                cell.font = company_font
        elif typ_radku == "AGENT":
            # Makléř - odsaď a přidej hyperlink
            jmeno_cell = ws.cell(row=row_idx, column=headers.index("jmeno_maklere") + 1 if "jmeno_maklere" in headers else 4)
            jmeno_cell.value = f"  → {jmeno_cell.value}"  # Odsazení

            # Hyperlink
            if profil_col:
                cell = ws.cell(row=row_idx, column=profil_col)
                url = cell.value
                if url and isinstance(url, str) and url.startswith("http"):
                    cell.hyperlink = url
                    cell.value = "Profil makléře"
                    cell.font = Font(color="0000FF", underline="single")

    # Šířka sloupců
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(output_path)

    # Spočítej statistiky
    companies_count = sum(1 for r in records if r.get("typ_radku") == "COMPANY")
    agents_count = sum(1 for r in records if r.get("typ_radku") == "AGENT")
    total_estates = sum(r.get("pocet_inzeratu", 0) for r in records if isinstance(r.get("pocet_inzeratu"), int))

    print(f"\n✅ Uloženo do: {output_path}")
    print(f"📊 Realitních kanceláří: {companies_count}")
    print(f"👤 Celkem makléřů: {agents_count}")
    print(f"🏠 Celkem inzerátů: {total_estates}")


def prompt_for_params():
    """Interaktivní výběr parametrů s podporou multiple selection."""
    print("\n" + "="*80)
    print("📋 INTERAKTIVNÍ VÝBĚR PARAMETRŮ")
    print("="*80)
    print()

    # Kategorie
    print("Typ nemovitosti:")
    print("  1 = Byty")
    print("  2 = Domy")
    print("  3 = Pozemky")
    print("  4 = Komerční")
    print("  5 = Ostatní")
    category_input = input("\nVyber typ nemovitosti (1-5, oddělené čárkou) [1]: ").strip() or "1"
    categories = [c.strip() for c in category_input.split(",")]
    category_main_list = [int(c) for c in categories if c.isdigit()]

    # Typy inzerátů
    print("\nTyp inzerátu:")
    print("  1 = Prodej")
    print("  2 = Pronájem")
    print("  3 = Dražby")
    type_input = input("\nVyber typ inzerátu (1-3, oddělené čárkou) [1]: ").strip() or "1"
    types = [t.strip() for t in type_input.split(",")]
    category_type_list = [int(t) for t in types if t.isdigit()]

    # Kraje
    print("\nKraj (volitelné):")
    print("  10 = Praha")
    print("  11 = Středočeský")
    print("  12 = Jihočeský")
    print("  13 = Plzeňský")
    print("  14 = Karlovarský")
    print("  15 = Ústecký")
    print("  16 = Liberecký")
    print("  17 = Královéhradecký")
    print("  18 = Pardubický")
    print("  19 = Vysočina")
    print("  20 = Jihomoravský")
    print("  21 = Olomoucký")
    print("  22 = Zlínský")
    print("  23 = Moravskoslezský")
    locality_input = input("\nVyber kraje (10-23, oddělené čárkami) nebo Enter pro celou ČR: ").strip()

    locality_list = None
    if locality_input:
        localities = [l.strip() for l in locality_input.split(",")]
        locality_list = [int(l) for l in localities if l.isdigit()]

    # Stránky
    print("\nPočet stránek:")
    pages_input = input("Max stránek (nebo 'all' pro všechny) [5]: ").strip() or "5"
    if pages_input.lower() == "all":
        max_pages = None
        full_scan = True
    else:
        max_pages = int(pages_input) if pages_input.isdigit() else 5
        full_scan = False

    print("\n" + "="*80)
    print(f"✅ Vybrané parametry:")
    print(f"   Typy nemovitostí: {category_main_list}")
    print(f"   Typy inzerátů: {category_type_list}")
    print(f"   Kraje: {locality_list or 'Celá ČR'}")
    print(f"   Stránek: {'VŠECHNY' if full_scan else max_pages}")
    print("="*80)
    print()

    return {
        "category_main_list": category_main_list,
        "category_type_list": category_type_list,
        "locality_list": locality_list,
        "max_pages": max_pages,
        "full_scan": full_scan,
    }


def merge_records(all_records):
    """Sloučí záznamy z více scrapování."""
    # Pro fast scraper jen spojíme všechny záznamy
    # (každý má unikátní company + makléř kombinaci)
    return all_records


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)

    parser.add_argument("--prompt", action="store_true", help="Interaktivní výběr parametrů")
    parser.add_argument("--category-main", type=int, default=1, help="1=Byty, 2=Domy, ...")
    parser.add_argument("--category-type", type=int, default=1, help="1=Prodej, 2=Pronájem, ...")
    parser.add_argument("--locality", type=int, help="10=Praha, 11=Středočeský, ...")
    parser.add_argument("--max-pages", type=int, default=5, help="Max stránek [5]")
    parser.add_argument("--full-scan", action="store_true", help="Všechny stránky")
    parser.add_argument("-o", "--output", help="Výstupní soubor")

    args = parser.parse_args()

    print("="*80)
    print("🚀 SUPER RYCHLÝ SCRAPER MAKLÉŘŮ (s company API)")
    print("="*80)
    print()

    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
    region_names = {
        10: "Praha", 11: "Středočeský", 12: "Jihočeský", 13: "Plzeňský",
        14: "Karlovarský", 15: "Ústecký", 16: "Liberecký", 17: "Královéhradecký",
        18: "Pardubický", 19: "Vysočina", 20: "Jihomoravský", 21: "Olomoucký",
        22: "Zlínský", 23: "Moravskoslezský"
    }

    try:
        scraper = SrealityScraper()

        if args.prompt:
            # Interaktivní mód
            params = prompt_for_params()

            # Vytvoř VŠECHNY kombinace najednou (pro deduplikaci!)
            combinations = []
            for category_main in params["category_main_list"]:
                for category_type in params["category_type_list"]:
                    localities = params["locality_list"] or [None]
                    for locality in localities:
                        combinations.append((category_main, category_type, locality))

            # Výpis kombinací
            print("\n" + "="*80)
            print(f"🎯 Celkem {len(combinations)} kombinací k zpracování:")
            for idx, (cat, typ, loc) in enumerate(combinations, 1):
                line = f"   {idx}. {category_names.get(cat)} / {type_names.get(typ)}"
                if loc:
                    line += f" / {region_names.get(loc)}"
                print(line)
            print("="*80)

            # Použij COMBINED funkci - automaticky deduplikuje companies!
            final_records = scrape_agents_fast_combined(
                scraper,
                combinations,
                params["max_pages"],
                params["full_scan"],
            )

        else:
            # Manuální parametry
            print("📋 Parametry:")
            print(f"   • Typ: {category_names.get(args.category_main, 'Neznámý')}")
            print(f"   • Inzerát: {type_names.get(args.category_type, 'Neznámý')}")
            print(f"   • Kraj: {region_names.get(args.locality, 'Celá ČR')}")
            print(f"   • Stránek: {'VŠECHNY' if args.full_scan else args.max_pages}")
            print()

            final_records = scrape_agents_fast(
                scraper,
                args.category_main,
                args.category_type,
                args.locality,
                args.max_pages,
                args.full_scan,
            )

        if final_records:
            output = args.output or f"data/makleri_fast_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_to_excel_hierarchical(final_records, output)
        else:
            print("⚠️  Žádná data")

    except KeyboardInterrupt:
        print("\n⚠️  Přerušeno")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Chyba: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    print("\n" + "="*80)
    print("✅ Hotovo!")
    print("="*80)


if __name__ == "__main__":
    main()
