#!/usr/bin/env python3
"""
🚀 JEDNODUCHÝ A RYCHLÝ scraper aktivních makléřů

Efektivní přístup:
- Projde inzeráty podle kategorie
- Agreguje data o maklérích PŘÍMO z inzerátů (bez dalších API volání!)
- Rychlé - zpracuje stránku za pár sekund

Výstup:
- Jméno, telefon, email, company
- Počet inzerátů podle typu
- URL profilu
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from scrapers.sreality import SrealityScraper


def scrape_agents_simple(
    scraper,
    category_main,
    category_type,
    locality_region_id,
    max_pages,
    full_scan,
):
    """Optimalizovaný scraping - agregace podle user_id + detail jen pro makléře bez kontaktů."""

    print(f"🔍 Scraping inzerátů pro získání makléřů...")

    if full_scan:
        max_pages = None

    limit = max_pages if max_pages is not None else None

    # Agregace dat o maklérích
    agents = defaultdict(lambda: {
        "user_id": None,
        "jmeno": None,
        "telefon": None,
        "email": None,
        "company": None,
        "kraj": None,
        "mesto": None,
        "inzeraty_breakdown": defaultdict(int),  # (category, type) -> count
        "total_count": 0,
        "sample_hash_id": None,  # Pro případné doplnění detailů
    })

    page = 1
    total_listings = 0

    # FÁZE 1: Projdi inzeráty a agreguj podle user_id
    while True:
        if limit is not None and page > limit:
            break

        params = {
            "category_main_cb": category_main,
            "category_type_cb": category_type,
            "page": page,
            "per_page": 60,
        }

        if locality_region_id is not None:
            params["locality_region_id"] = locality_region_id

        payload = scraper._request(scraper._config.api_url, params=params)
        if not payload:
            print(f"⚠️  Chyba při stahování stránky {page}")
            break

        estates = payload.get("_embedded", {}).get("estates", [])
        if not estates:
            break

        print(f"   Stránka {page}: {len(estates)} inzerátů")

        # Debug counters
        debug_no_embedded = 0
        debug_no_seller_broker = 0
        debug_no_user_id = 0

        # Zpracuj každý inzerát
        for estate in estates:
            total_listings += 1

            # Získej user_id makléře z _links nebo _embedded
            user_id = None
            embedded = estate.get("_embedded", {})

            if not embedded:
                debug_no_embedded += 1
                continue

            # Zkus různé zdroje user_id
            seller = embedded.get("seller", {})
            broker = embedded.get("broker", {})

            if not seller and not broker:
                debug_no_seller_broker += 1
                # DEBUG: Vypsat dostupné klíče v embedded
                if total_listings <= 3:  # Jen pro první 3 inzeráty
                    print(f"   DEBUG inzerát #{total_listings}: _embedded keys = {list(embedded.keys())}")
                continue

            user_id = (
                seller.get("user_id")
                or seller.get("id")
                or broker.get("user_id")
                or broker.get("id")
            )

            if not user_id:
                debug_no_user_id += 1
                # DEBUG: Vypsat strukturu sellera/brokera
                if total_listings <= 3:
                    print(f"   DEBUG inzerát #{total_listings}:")
                    if seller:
                        print(f"      seller keys: {list(seller.keys())}")
                    if broker:
                        print(f"      broker keys: {list(broker.keys())}")
                continue

            user_id = str(user_id)
            agent = agents[user_id]

            # První výskyt - ulož základní info
            if agent["user_id"] is None:
                agent["user_id"] = user_id
                agent["sample_hash_id"] = estate.get("hash_id")  # Uchováme pro případný detail
                agent["jmeno"] = (
                    seller.get("user_name")
                    or seller.get("name")
                    or broker.get("user_name")
                    or broker.get("name")
                    or "Neznámý makléř"
                )
                agent["company"] = (
                    seller.get("company_name")
                    or seller.get("company", {}).get("name") if isinstance(seller.get("company"), dict) else None
                    or embedded.get("company", {}).get("name")
                )

                # Telefon a email z embedded (pokud jsou dostupné)
                phones = embedded.get("phones", [])
                if phones and isinstance(phones, list):
                    for phone in phones:
                        if isinstance(phone, dict):
                            agent["telefon"] = phone.get("number") or phone.get("value")
                            if agent["telefon"]:
                                break
                        elif isinstance(phone, str):
                            agent["telefon"] = phone
                            break

                emails = embedded.get("emails", [])
                if emails and isinstance(emails, list):
                    for email in emails:
                        if isinstance(email, dict):
                            agent["email"] = email.get("value") or email.get("email")
                            if agent["email"]:
                                break
                        elif isinstance(email, str):
                            agent["email"] = email
                            break

                # Lokalita
                locality = estate.get("locality", "")
                if locality:
                    parts = [p.strip() for p in locality.split(",")]
                    if parts:
                        agent["mesto"] = parts[0]
                        if len(parts) > 1:
                            agent["kraj"] = parts[-1]

            # Doplň kontakty pokud chybí
            if not agent["telefon"]:
                phones = embedded.get("phones", [])
                if phones and isinstance(phones, list):
                    for phone in phones:
                        if isinstance(phone, dict):
                            agent["telefon"] = phone.get("number") or phone.get("value")
                            if agent["telefon"]:
                                break

            if not agent["email"]:
                emails = embedded.get("emails", [])
                if emails and isinstance(emails, list):
                    for email in emails:
                        if isinstance(email, dict):
                            agent["email"] = email.get("value") or email.get("email")
                            if agent["email"]:
                                break

            # Spočítej typ inzerátu
            seo = estate.get("seo", {}) if isinstance(estate.get("seo"), dict) else {}
            cat_main = seo.get("category_main_cb") or category_main
            cat_type = seo.get("category_type_cb") or category_type

            key = (cat_main, cat_type)
            agent["inzeraty_breakdown"][key] += 1
            agent["total_count"] += 1

        # DEBUG: Vypsat statistiku
        if debug_no_embedded > 0 or debug_no_seller_broker > 0 or debug_no_user_id > 0:
            print(f"   ⚠️  DEBUG statistika:")
            print(f"      Inzerátů bez _embedded: {debug_no_embedded}")
            print(f"      Inzerátů bez seller/broker: {debug_no_seller_broker}")
            print(f"      Inzerátů s seller/broker ale bez user_id: {debug_no_user_id}")

        # Kontrola konce
        result_size = payload.get("result_size", 0)
        if (page * 60) >= result_size:
            break

        page += 1
        scraper._delay()

    print(f"\n✅ Zpracováno {total_listings} inzerátů")
    print(f"✅ Nalezeno {len(agents)} unikátních makléřů")

    # FÁZE 2: Doplň kontakty pro makléře, kteří je nemají
    agents_without_contacts = [
        (user_id, agent) for user_id, agent in agents.items()
        if not agent["telefon"] or not agent["email"]
    ]

    if agents_without_contacts:
        print(f"\n🔍 Doplňuji kontakty pro {len(agents_without_contacts)} makléřů bez telefonu/emailu...")

        for idx, (user_id, agent) in enumerate(agents_without_contacts, 1):
            hash_id = agent.get("sample_hash_id")
            if not hash_id:
                continue

            # Stáhni detail JEDNOHO inzerátu
            detail_url = f"{scraper._config.base_url}/api/cs/v2/estates/{hash_id}"
            detail = scraper._request(detail_url)

            if detail:
                embedded = detail.get("_embedded", {})

                # Doplň telefon
                if not agent["telefon"]:
                    phones = embedded.get("phones", [])
                    if phones and isinstance(phones, list):
                        for phone in phones:
                            if isinstance(phone, dict):
                                agent["telefon"] = phone.get("number") or phone.get("value")
                                if agent["telefon"]:
                                    break

                # Doplň email
                if not agent["email"]:
                    emails = embedded.get("emails", [])
                    if emails and isinstance(emails, list):
                        for email in emails:
                            if isinstance(email, dict):
                                agent["email"] = email.get("value") or email.get("email")
                                if agent["email"]:
                                    break

                scraper._delay()

                if idx % 10 == 0:
                    print(f"   Doplněno {idx}/{len(agents_without_contacts)}...")

        print(f"✅ Doplněno kontaktů")

    # Převeď na finální formát
    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}

    final_records = []
    for user_id, agent in agents.items():
        # Vytvoř rozložení inzerátů
        breakdown_items = []
        for (cat, typ), count in sorted(agent["inzeraty_breakdown"].items(), key=lambda x: -x[1]):
            cat_name = category_names.get(cat, f"Kategorie {cat}")
            typ_name = type_names.get(typ, f"Typ {typ}")
            breakdown_items.append(f"{cat_name}/{typ_name}: {count}")

        rozlozeni = ", ".join(breakdown_items) if breakdown_items else "Neznámé"

        # URL profilu
        profil_url = f"https://www.sreality.cz/makler/{user_id}"

        final_records.append({
            "zdroj": "Sreality.cz",
            "jmeno_maklere": agent["jmeno"],
            "telefon": agent["telefon"],
            "email": agent["email"],
            "realitni_kancelar": agent["company"],
            "kraj": agent["kraj"],
            "mesto": agent["mesto"],
            "profil_url": profil_url,
            "pocet_inzeratu": agent["total_count"],
            "rozlozeni_inzeratu": rozlozeni,
        })

    return final_records


def prompt_for_params():
    """Interaktivní výběr parametrů s podporou multiple selection."""
    print("\n" + "="*80)
    print("📋 INTERAKTIVNÍ VÝBĚR PARAMETRŮ")
    print("="*80)
    print()

    # Kategorie
    print("Typ nemovitosti:")
    print("  1 = Byty")
    print("  2 = Domy")
    print("  3 = Pozemky")
    print("  4 = Komerční")
    print("  5 = Ostatní")
    category_input = input("\nVyber typ nemovitosti (1-5) [1]: ").strip() or "1"
    categories = [c.strip() for c in category_input.split(",")]
    category_main_list = [int(c) for c in categories if c.isdigit()]

    # Typy inzerátů
    print("\nTyp inzerátu:")
    print("  1 = Prodej")
    print("  2 = Pronájem")
    print("  3 = Dražby")
    type_input = input("\nVyber typ inzerátu (1-3) [1]: ").strip() or "1"
    types = [t.strip() for t in type_input.split(",")]
    category_type_list = [int(t) for t in types if t.isdigit()]

    # Kraje
    print("\nKraj (volitelné):")
    print("  10 = Praha")
    print("  11 = Středočeský")
    print("  12 = Jihočeský")
    print("  13 = Plzeňský")
    print("  14 = Karlovarský")
    print("  15 = Ústecký")
    print("  16 = Liberecký")
    print("  17 = Královéhradecký")
    print("  18 = Pardubický")
    print("  19 = Vysočina")
    print("  20 = Jihomoravský")
    print("  21 = Olomoucký")
    print("  22 = Zlínský")
    print("  23 = Moravskoslezský")
    locality_input = input("\nVyber kraje (10-23, oddělené čárkami) nebo Enter pro celou ČR: ").strip()

    locality_list = None
    if locality_input:
        localities = [l.strip() for l in locality_input.split(",")]
        locality_list = [int(l) for l in localities if l.isdigit()]

    # Stránky
    print("\nPočet stránek:")
    pages_input = input("Max stránek (nebo 'all' pro všechny) [5]: ").strip() or "5"
    if pages_input.lower() == "all":
        max_pages = None
        full_scan = True
    else:
        max_pages = int(pages_input) if pages_input.isdigit() else 5
        full_scan = False

    print("\n" + "="*80)
    print(f"✅ Vybrané parametry:")
    print(f"   Typy nemovitostí: {category_main_list}")
    print(f"   Typy inzerátů: {category_type_list}")
    print(f"   Kraje: {locality_list or 'Celá ČR'}")
    print(f"   Stránek: {'VŠECHNY' if full_scan else max_pages}")
    print("="*80)
    print()

    return {
        "category_main_list": category_main_list,
        "category_type_list": category_type_list,
        "locality_list": locality_list,
        "max_pages": max_pages,
        "full_scan": full_scan,
    }


def save_to_excel(records, output_path):
    """Uloží do Excelu s formátováním."""
    if not records:
        print("⚠️  Žádné záznamy")
        return

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(records)
    df = df.sort_values(by="pocet_inzeratu", ascending=False)
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Formátování
    wb = load_workbook(output_path)
    ws = wb.active

    # Hyperlinky
    headers = [cell.value for cell in ws[1]]
    profil_col = None
    for idx, header in enumerate(headers, 1):
        if header == "profil_url":
            profil_col = idx
            break

    if profil_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=profil_col)
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.value = "Profil makléře"
                cell.font = Font(color="0000FF", underline="single")

    # Šířka sloupců
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(output_path)

    print(f"\n✅ Uloženo do: {output_path}")
    print(f"📊 Celkem makléřů: {len(df)}")
    if "pocet_inzeratu" in df.columns:
        print(f"🏠 Celkem inzerátů: {df['pocet_inzeratu'].sum()}")


def merge_agents(all_records):
    """Sloučí záznamy makléřů z více scrapování podle user_id."""
    from collections import defaultdict
    import re

    merged = {}

    for record in all_records:
        # Klíč podle jména + telefon/email
        key = (
            record.get("jmeno_maklere", "").strip().lower(),
            record.get("telefon", "").strip(),
            record.get("email", "").strip(),
        )

        # Pokud je klíč prázdný, zkus user_id z URL
        if not any(key):
            url = record.get("profil_url", "")
            match = re.search(r"/makler/(\d+)", url)
            if match:
                user_id = match.group(1)
                key = (user_id, "", "")

        key_str = str(key)

        if key_str not in merged:
            merged[key_str] = record.copy()
        else:
            # Slouč inzeráty
            existing = merged[key_str]

            # Doplň chybějící kontakty
            if not existing.get("telefon") and record.get("telefon"):
                existing["telefon"] = record["telefon"]
            if not existing.get("email") and record.get("email"):
                existing["email"] = record["email"]
            if not existing.get("realitni_kancelar") and record.get("realitni_kancelar"):
                existing["realitni_kancelar"] = record["realitni_kancelar"]

            # Agreguj počty
            existing["pocet_inzeratu"] += record.get("pocet_inzeratu", 0)

            # Slouč rozložení
            breakdown1 = existing.get("rozlozeni_inzeratu", "")
            breakdown2 = record.get("rozlozeni_inzeratu", "")

            # Parse a agregguj
            breakdown_dict = defaultdict(int)

            for breakdown_text in [breakdown1, breakdown2]:
                if not breakdown_text or breakdown_text == "Neznámé":
                    continue
                for part in breakdown_text.split(", "):
                    if ":" in part:
                        key_part, count_part = part.rsplit(":", 1)
                        try:
                            count = int(count_part.strip())
                            breakdown_dict[key_part.strip()] += count
                        except:
                            pass

            # Vytvoř nové rozložení
            breakdown_items = [f"{k}: {v}" for k, v in sorted(breakdown_dict.items(), key=lambda x: -x[1])]
            existing["rozlozeni_inzeratu"] = ", ".join(breakdown_items) if breakdown_items else "Neznámé"

    return list(merged.values())


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)

    parser.add_argument("--prompt", action="store_true", help="Interaktivní výběr parametrů")
    parser.add_argument("--category-main", type=int, default=1, help="1=Byty, 2=Domy, ...")
    parser.add_argument("--category-type", type=int, default=1, help="1=Prodej, 2=Pronájem, ...")
    parser.add_argument("--locality", type=int, help="10=Praha, 11=Středočeský, ...")
    parser.add_argument("--max-pages", type=int, default=5, help="Max stránek [5]")
    parser.add_argument("--full-scan", action="store_true", help="Všechny stránky")
    parser.add_argument("-o", "--output", help="Výstupní soubor")

    args = parser.parse_args()

    print("="*80)
    print("🚀 RYCHLÝ SCRAPER MAKLÉŘŮ (optimalizovaný)")
    print("="*80)
    print()

    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
    region_names = {
        10: "Praha", 11: "Středočeský", 12: "Jihočeský", 13: "Plzeňský",
        14: "Karlovarský", 15: "Ústecký", 16: "Liberecký", 17: "Královéhradecký",
        18: "Pardubický", 19: "Vysočina", 20: "Jihomoravský", 21: "Olomoucký",
        22: "Zlínský", 23: "Moravskoslezský"
    }

    try:
        scraper = SrealityScraper()

        if args.prompt:
            # Interaktivní mód
            params = prompt_for_params()

            # Vytvoř kombinace parametrů
            all_records = []

            for category_main in params["category_main_list"]:
                for category_type in params["category_type_list"]:
                    localities = params["locality_list"] or [None]

                    for locality in localities:
                        print("\n" + "="*80)
                        print(f"🔍 Scraping: {category_names.get(category_main)} / {type_names.get(category_type)}")
                        if locality:
                            print(f"   Kraj: {region_names.get(locality)}")
                        print("="*80)

                        records = scrape_agents_simple(
                            scraper,
                            category_main,
                            category_type,
                            locality,
                            params["max_pages"],
                            params["full_scan"],
                        )

                        all_records.extend(records)

            # Slouč duplicity
            print("\n" + "="*80)
            print("🔄 Slučování duplicitních makléřů...")
            print("="*80)
            final_records = merge_agents(all_records)
            print(f"✅ Sloučeno z {len(all_records)} záznamů na {len(final_records)} unikátních makléřů")

        else:
            # Manuální parametry
            print("📋 Parametry:")
            print(f"   • Typ: {category_names.get(args.category_main, 'Neznámý')}")
            print(f"   • Inzerát: {type_names.get(args.category_type, 'Neznámý')}")
            print(f"   • Kraj: {region_names.get(args.locality, 'Celá ČR')}")
            print(f"   • Stránek: {'VŠECHNY' if args.full_scan else args.max_pages}")
            print()

            final_records = scrape_agents_simple(
                scraper,
                args.category_main,
                args.category_type,
                args.locality,
                args.max_pages,
                args.full_scan,
            )

        if final_records:
            output = args.output or f"data/makleri_simple_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_to_excel(final_records, output)
        else:
            print("⚠️  Žádná data")

    except KeyboardInterrupt:
        print("\n⚠️  Přerušeno")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Chyba: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    print("\n" + "="*80)
    print("✅ Hotovo!")
    print("="*80)


if __name__ == "__main__":
    main()
