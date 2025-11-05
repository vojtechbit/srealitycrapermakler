#!/usr/bin/env python3
"""
🎯 EFEKTIVNÍ SCRAPER: Aktivní makléři s kompletními profily

Tento scraper kombinuje oba přístupy:
1. Najde aktivní makléře podle kategorie/kraje (rychlé)
2. Pro každého získá VŠECHNY inzeráty a kompletní profil (přesné)

Výsledek: Aktivní makléři s přesným počtem inzerátů a správnými URL profilu.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from scrapers.sreality import SrealityScraper


def prompt_for_params() -> dict:
    """Interaktivně se zeptá na parametry."""
    print("\n" + "="*60)
    print("NASTAVENÍ PARAMETRŮ")
    print("="*60)

    # Kategorie nemovitosti
    print("\n📋 Typ nemovitosti (můžeš vybrat víc oddělených čárkou):")
    print("  1 - Byty")
    print("  2 - Domy")
    print("  3 - Pozemky")
    print("  4 - Komerční")
    print("  5 - Ostatní")
    print("  Příklad: '1,2' = Byty a Domy")

    while True:
        category_input = input("Vyber typ nemovitosti (1-5) [1]: ").strip() or "1"
        categories = [c.strip() for c in category_input.split(",")]
        if all(c in ["1", "2", "3", "4", "5"] for c in categories):
            category_main_list = [int(c) for c in categories]
            break
        print("❌ Neplatná volba, zadej čísla 1-5 oddělená čárkou (např. '1,2')")

    # Typ inzerátu
    print("\n📋 Typ inzerátu (můžeš vybrat víc oddělených čárkou):")
    print("  1 - Prodej")
    print("  2 - Pronájem")
    print("  3 - Dražby")
    print("  Příklad: '1,2' = Prodej a Pronájem")

    while True:
        type_input = input("Vyber typ inzerátu (1-3) [1]: ").strip() or "1"
        types = [t.strip() for t in type_input.split(",")]
        if all(t in ["1", "2", "3"] for t in types):
            category_type_list = [int(t) for t in types]
            break
        print("❌ Neplatná volba, zadej čísla 1-3 oddělená čárkou (např. '1,2')")

    # Kraj
    print("\n📋 Kraj (můžeš vybrat víc oddělených čárkou):")
    print("  10 - Praha")
    print("  11 - Středočeský")
    print("  12 - Jihočeský")
    print("  13 - Plzeňský")
    print("  14 - Karlovarský")
    print("  15 - Ústecký")
    print("  16 - Liberecký")
    print("  17 - Královéhradecký")
    print("  18 - Pardubický")
    print("  19 - Vysočina")
    print("  20 - Jihomoravský")
    print("  21 - Olomoucký")
    print("  22 - Zlínský")
    print("  23 - Moravskoslezský")
    print("  Příklad: '10,20' = Praha a Jihomoravský")

    locality_input = input("Vyber kraje (10-23) oddělené čárkou nebo Enter pro celou ČR: ").strip()
    if locality_input:
        localities = [loc.strip() for loc in locality_input.split(",")]
        try:
            locality_list = [int(loc) for loc in localities if loc]
            # Validace
            if not all(10 <= loc <= 23 for loc in locality_list):
                print("⚠️  Některé kódy krajů jsou mimo rozsah 10-23, budu je ignorovat")
                locality_list = [loc for loc in locality_list if 10 <= loc <= 23]
        except ValueError:
            print("⚠️  Neplatné číslo kraje, použiji celou ČR")
            locality_list = [None]
    else:
        locality_list = [None]  # Celá ČR

    # Maximální počet stránek
    print("\n📋 Rozsah scrapování:")
    print("  0 - VŠECHNY stránky (může trvat hodiny!)")
    print("  1-10 - Rychlý test")
    print("  20-50 - Střední rozsah")
    print("  100+ - Velký rozsah")

    while True:
        max_pages = input("Maximální počet stránek [5]: ").strip() or "5"
        try:
            max_pages = int(max_pages)
            if max_pages < 0:
                print("❌ Zadej kladné číslo nebo 0")
                continue
            break
        except ValueError:
            print("❌ Zadej platné číslo")

    # Full scan
    full_scan = False
    if max_pages == 0:
        full_scan = True
        max_pages = None

    # Detaily
    print("\n📋 Stahovat detaily inzerátů?")
    print("  y - Ano (přesnější kontakty, ale POMALEJŠÍ)")
    print("  n - Ne (rychlejší, ale méně kontaktů)")

    fetch_details = input("Stahovat detaily? (y/n) [y]: ").strip().lower() or "y"
    fetch_details = fetch_details == "y"

    # Vypočítej celkový počet kombinací
    total_combinations = len(category_main_list) * len(category_type_list) * len(locality_list)

    print("\n" + "="*60)
    print("SOUHRN:")
    print("="*60)
    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
    region_names = {
        10: "Praha", 11: "Středočeský", 12: "Jihočeský", 13: "Plzeňský",
        14: "Karlovarský", 15: "Ústecký", 16: "Liberecký", 17: "Královéhradecký",
        18: "Pardubický", 19: "Vysočina", 20: "Jihomoravský", 21: "Olomoucký",
        22: "Zlínský", 23: "Moravskoslezský"
    }

    category_names_str = ", ".join(category_names.get(c, 'Neznámý') for c in category_main_list)
    type_names_str = ", ".join(type_names.get(t, 'Neznámý') for t in category_type_list)
    region_names_str = ", ".join(region_names.get(loc, 'Celá ČR') for loc in locality_list)

    print(f"📌 Typ nemovitosti: {category_names_str}")
    print(f"📌 Typ inzerátu: {type_names_str}")
    print(f"📌 Kraje: {region_names_str}")
    print(f"📌 Max. stránek: {'VŠECHNY' if full_scan else max_pages}")
    print(f"📌 Detaily: {'Ano' if fetch_details else 'Ne'}")
    print(f"\n⚠️  Celkem kombinací k scrapování: {total_combinations}")
    if total_combinations > 1:
        print(f"    (Scraper poběží {total_combinations}x a výsledky sloučí)")
    print("="*60)

    confirm = input("\nPokračovat? (y/n) [y]: ").strip().lower() or "y"
    if confirm != "y":
        print("❌ Zrušeno uživatelem")
        sys.exit(0)

    return {
        "category_main_list": category_main_list,
        "category_type_list": category_type_list,
        "locality_list": locality_list,
        "max_pages": max_pages,
        "full_scan": full_scan,
        "fetch_details": fetch_details,
    }


def save_to_excel_with_formatting(records: list, output_path: str) -> None:
    """Uloží data do Excelu s hyperlinky a formátováním."""
    if not records:
        print("⚠️  Žádné záznamy k uložení.")
        return

    # Vytvoř složku
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    # Převeď na DataFrame a seřaď
    df = pd.DataFrame(records)
    if "pocet_inzeratu" in df.columns:
        df = df.sort_values(by="pocet_inzeratu", ascending=False)

    # Ulož do Excelu
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Přidej hyperlinky a formátování
    wb = load_workbook(output_path)
    ws = wb.active

    # Najdi sloupce s linky
    headers = [cell.value for cell in ws[1]]
    link_columns = []

    for idx, header in enumerate(headers, 1):
        if header in ["profil_url"]:
            link_columns.append((idx, header))

    # Přidej hyperlinky
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in link_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value

            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.value = "Profil makléře"
                cell.font = Font(color="0000FF", underline="single")

    # Automatická šířka sloupců
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

    print(f"\n✅ Data uložena do: {output_path}")
    print(f"📊 Celkem makléřů: {len(df)}")

    if "pocet_inzeratu" in df.columns:
        total_listings = df["pocet_inzeratu"].sum()
        print(f"🏠 Celkem inzerátů: {total_listings}")


def main():
    """Hlavní funkce."""
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "--category-main",
        type=int,
        default=1,
        help="Typ nemovitosti (1=Byty, 2=Domy, 3=Pozemky, 4=Komerční, 5=Ostatní) [výchozí: 1]",
    )

    parser.add_argument(
        "--category-type",
        type=int,
        default=1,
        help="Typ inzerátu (1=Prodej, 2=Pronájem, 3=Dražby) [výchozí: 1]",
    )

    parser.add_argument(
        "--locality",
        type=int,
        help="ID kraje (10=Praha, 11=Středočeský, ..., 23=Moravskoslezský) [výchozí: celá ČR]",
    )

    parser.add_argument(
        "--max-pages",
        type=int,
        default=5,
        help="Maximální počet stránek pro hledání makléřů [výchozí: 5]",
    )

    parser.add_argument(
        "--full-scan",
        action="store_true",
        help="Projít VŠECHNY stránky (může trvat hodiny)",
    )

    parser.add_argument(
        "--no-details",
        action="store_true",
        help="Nestahovat detaily inzerátů (rychlejší, ale méně přesné kontakty)",
    )

    parser.add_argument(
        "-o", "--output",
        help="Cesta k výstupnímu souboru [výchozí: data/active_agents_TIMESTAMP.xlsx]",
    )

    parser.add_argument(
        "--prompt",
        action="store_true",
        help="Interaktivní mód - zeptá se na všechny parametry",
    )

    args = parser.parse_args()

    print("="*80)
    print("🎯 SCRAPER AKTIVNÍCH MAKLÉŘŮ S KOMPLETNÍMI PROFILY")
    print("="*80)
    print()

    if not args.prompt:
        print("💡 Tip: Použij --prompt pro interaktivní mód")
        print("   nebo přečti si README_ACTIVE_AGENTS.md")
        print()

    category_names = {1: "Byty", 2: "Domy", 3: "Pozemky", 4: "Komerční", 5: "Ostatní"}
    type_names = {1: "Prodej", 2: "Pronájem", 3: "Dražby"}
    region_names = {
        10: "Praha", 11: "Středočeský", 12: "Jihočeský", 13: "Plzeňský",
        14: "Karlovarský", 15: "Ústecký", 16: "Liberecký", 17: "Královéhradecký",
        18: "Pardubický", 19: "Vysočina", 20: "Jihomoravský", 21: "Olomoucký",
        22: "Zlínský", 23: "Moravskoslezský"
    }

    # Interaktivní mód nebo parametry z příkazové řádky?
    if args.prompt:
        params = prompt_for_params()
        category_main_list = params["category_main_list"]
        category_type_list = params["category_type_list"]
        locality_list = params["locality_list"]
        max_pages = params["max_pages"]
        full_scan = params["full_scan"]
        fetch_details = params["fetch_details"]
    else:
        # Jednoduchý režim - z argumentů
        category_main_list = [args.category_main]
        category_type_list = [args.category_type]
        locality_list = [args.locality]
        max_pages = args.max_pages
        full_scan = args.full_scan
        fetch_details = not args.no_details

        print("📋 Parametry:")
        print(f"   • Typ nemovitosti: {category_names.get(args.category_main, 'Neznámý')}")
        print(f"   • Typ inzerátu: {type_names.get(args.category_type, 'Neznámý')}")
        print(f"   • Kraj: {region_names.get(args.locality, 'Celá ČR')}")
        print(f"   • Max. stránek: {'VŠECHNY' if full_scan else max_pages}")
        print(f"   • Detaily: {'Ano' if fetch_details else 'Ne'}")
        print()

    print("\n⏳ Spouštím scraping...")
    print("   Fáze 1: Najdu aktivní makléře podle kategorie")
    print("   Fáze 2: Pro každého získám všechny inzeráty a profil")
    print()

    try:
        scraper = SrealityScraper()
        all_records = []

        total_combinations = len(category_main_list) * len(category_type_list) * len(locality_list)
        current_combo = 0

        # Projdi všechny kombinace
        for category_main in category_main_list:
            for category_type in category_type_list:
                for locality in locality_list:
                    current_combo += 1

                    if total_combinations > 1:
                        print(f"\n{'='*60}")
                        print(f"Kombinace {current_combo}/{total_combinations}")
                        print(f"  • Typ: {category_names.get(category_main, 'Neznámý')}")
                        print(f"  • Inzerát: {type_names.get(category_type, 'Neznámý')}")
                        print(f"  • Kraj: {region_names.get(locality, 'Celá ČR')}")
                        print('='*60)

                    result = scraper.scrape_active_agents_full_profiles(
                        category_main=category_main,
                        category_type=category_type,
                        locality_region_id=locality,
                        max_pages=max_pages,
                        full_scan=full_scan,
                        fetch_details=fetch_details,
                    )

                    if result.errors:
                        print("\n⚠️  Chyby:")
                        for error in result.errors:
                            print(f"   • {error}")

                    if result.records:
                        print(f"✅ Získáno {len(result.records)} makléřů z této kombinace")
                        all_records.extend(result.records)

        # Deduplikuj makléře
        if all_records:
            # Deduplikace podle jména + telefon + email + kancelář
            unique_records = {}
            for record in all_records:
                key = (
                    record.get("jmeno_maklere"),
                    record.get("telefon"),
                    record.get("email"),
                    record.get("realitni_kancelar")
                )
                key_str = "|".join(str(v) if v else "" for v in key)

                if key_str not in unique_records:
                    unique_records[key_str] = record

            final_records = list(unique_records.values())
            print(f"\n✅ Celkem {len(final_records)} unikátních makléřů (po deduplikaci)")

            # Výstupní soubor
            if args.output:
                output_path = args.output
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"data/active_agents_{timestamp}.xlsx"

            save_to_excel_with_formatting(final_records, output_path)

        else:
            print("\n⚠️  Nepodařilo se získat žádná data.")

    except KeyboardInterrupt:
        print("\n\n⚠️  Přerušeno uživatelem (Ctrl+C)")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Chyba: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    print("\n" + "="*80)
    print("✅ Hotovo!")
    print("="*80)


if __name__ == "__main__":
    main()
